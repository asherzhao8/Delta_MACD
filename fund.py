import pandas as pd
import re
import os
from openpyxl import load_workbook

path = 'd:/估值表/'
files = os.listdir(path)   #未清洗的文件名
filenames = []
for filename in files:
    if bool(re.search('\D', filename)) == False:
        filenames.append(filename)
date_sort = sorted(list(map(int,filenames)))
today = str(date_sort[-1])
today_str = today[:4] + '-' + today[4:6] + '-' + today[6:]
yesterday = str(date_sort[-2])
yesterday_str = yesterday[:4] + '-' + yesterday[4:6] + '-' + yesterday[6:]

try:
    fund1 = pd.read_excel(path + today + '/玉泉895号_' + yesterday_str + '.xls', skiprows = 3)
    fund1.set_index('科目代码', inplace = True)
except:
    print('无玉泉895号')
try:
    fund2 = pd.read_excel(path + today + '/证券投资基金估值表_信达澳银基金-战略2号单一资产管理计划_' + yesterday_str + '.xls', skiprows = 3)
    fund2.set_index('科目代码', inplace = True)
except:
    print('无战略2号')
try:
    fund3 = pd.read_excel(path + today + '/证券投资基金估值表_信达澳银基金-战略3号单一资产管理计划_' + yesterday_str + '.xls', skiprows = 3)
    fund3.set_index('科目代码', inplace = True)
except:
    print('无战略3号')
try:
    fund4 = pd.read_excel(path + today + '/证券投资基金估值表_信达澳银基金-战略4号单一资产管理计划_' + yesterday_str + '.xls', skiprows = 3)
    fund4.set_index('科目代码', inplace = True)
except:
    print('无战略4号')
try:
    fund5 = pd.read_excel(path + today + '/证券投资基金估值表_信达澳银基金-战略1号单一资产管理计划_' + yesterday_str + '.xls', skiprows = 3)
    fund5.set_index('科目代码', inplace = True)
except:
    print('无战略1号')

wb = load_workbook(path + yesterday + '/金融产品部交易日报' + yesterday + '.xlsx')
sheetnames = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheetnames[0])
max_row = sheet.max_row
max_cow = sheet.max_column

# 遍历行列锁定需要修改的位置
for i in range(max_row):
    i += 1 #表格从1开始
    try:
        if sheet.cell(i,4).value == '财通基金玉泉895号单一资产管理计划':
            for j in range(max_cow):
                j += 1
                if sheet.cell(2,j).value == '持仓数量\n（必填）':
                    sheet.cell(i,j).value = fund1.at['实收资本金额', '数量']
                if sheet.cell(2,j).value == '投入成本':
                    sheet.cell(i,j).value = round(fund1.at['基金资产净值:', '成本'] / 1e6, 0) * 1e6
                if sheet.cell(2,j).value == '单位净值':
                    sheet.cell(i,j).value = fund1.at['基金单位净值：', '市值']
                if sheet.cell(2,j).value == '浮动盈利':
                    sheet.cell(i,j).value = fund1.at['基金资产净值:', '市值'] - fund1.at['基金资产净值:', '成本']
    except:
        pass
    
    try:                
        if sheet.cell(i,4).value == '信达澳银基金-战略2号单一资产管理计划':
            for j in range(max_cow):
                j += 1
                if sheet.cell(2,j).value == '持仓数量\n（必填）':
                    sheet.cell(i,j).value = fund2.at['实收资本金额', '数量']
                if sheet.cell(2,j).value == '投入成本':
                    sheet.cell(i,j).value = round(fund2.at['基金资产净值:', '成本'] / 1e6, 0) * 1e6
                if sheet.cell(2,j).value == '单位净值':
                    sheet.cell(i,j).value = fund2.at['基金单位净值：', '市值']
                if sheet.cell(2,j).value == '浮动盈利':
                    sheet.cell(i,j).value = fund2.at['基金资产净值:', '市值'] - fund2.at['基金资产净值:', '成本']
    except:
        pass
                
    try: 
        if sheet.cell(i,4).value == '信达澳银基金-战略3号单一资产管理计划':
            for j in range(max_cow):
                j += 1
                if sheet.cell(2,j).value == '持仓数量\n（必填）':
                    sheet.cell(i,j).value = fund3.at['实收资本金额', '数量']
                if sheet.cell(2,j).value == '投入成本':
                    sheet.cell(i,j).value = round(fund3.at['基金资产净值:', '成本'] / 1e6, 0) * 1e6
                if sheet.cell(2,j).value == '单位净值':
                    sheet.cell(i,j).value = fund3.at['基金单位净值：', '市值']
                if sheet.cell(2,j).value == '浮动盈利':
                    sheet.cell(i,j).value = fund3.at['基金资产净值:', '市值'] - fund3.at['基金资产净值:', '成本']
    except:
        pass
                
    try: 
        if sheet.cell(i,4).value == '信达澳银基金-战略4号单一资产管理计划':
            for j in range(max_cow):
                j += 1
                if sheet.cell(2,j).value == '持仓数量\n（必填）':
                    sheet.cell(i,j).value = fund4.at['实收资本金额', '数量']
                if sheet.cell(2,j).value == '投入成本':
                    sheet.cell(i,j).value = round(fund4.at['基金资产净值:', '成本'] / 1e6, 0) * 1e6
                if sheet.cell(2,j).value == '单位净值':
                    sheet.cell(i,j).value = fund4.at['基金单位净值：', '市值']
                if sheet.cell(2,j).value == '浮动盈利':
                    sheet.cell(i,j).value = fund4.at['基金资产净值:', '市值'] - fund4.at['基金资产净值:', '成本']
    except:
        pass
             
    try: 
        if sheet.cell(i,4).value == '信达澳银基金-战略1号单一资产管理计划':
            for j in range(max_cow):
                j += 1
                if sheet.cell(2,j).value == '持仓数量\n（必填）':
                    sheet.cell(i,j).value = fund5.at['实收资本金额', '数量']
                if sheet.cell(2,j).value == '投入成本':
                    sheet.cell(i,j).value = round(fund5.at['基金资产净值:', '成本'] / 1e6, 0) * 1e6
                if sheet.cell(2,j).value == '单位净值':
                    sheet.cell(i,j).value = fund5.at['基金单位净值：', '市值']
                if sheet.cell(2,j).value == '浮动盈利':
                    sheet.cell(i,j).value = fund5.at['基金资产净值:', '市值'] - fund5.at['基金资产净值:', '成本']
    except:
        pass
                
wb.save(path + today + '/金融产品部交易日报' + today + '.xlsx')